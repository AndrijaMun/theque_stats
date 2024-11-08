import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt

# Create a new directory for the generated files
output_dir = 'analyzed_data_output'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Connect to the SQLite database
conn = sqlite3.connect('theque.db')

# Retrieve data from the tables
orders = pd.read_sql_query("SELECT * FROM Orders", conn)
order_items = pd.read_sql_query("SELECT * FROM OrderItems", conn)
items = pd.read_sql_query("SELECT * FROM Items", conn)
cashiers = pd.read_sql_query("SELECT * FROM Cashiers", conn)
order_types = pd.read_sql_query("SELECT * FROM OrderTypes", conn)
payment_types = pd.read_sql_query("SELECT * FROM PaymentTypes", conn)
addons = pd.read_sql_query("SELECT * FROM AddOns", conn)
item_addons = pd.read_sql_query("SELECT * FROM ItemAddOns", conn)
item_flavours = pd.read_sql_query("SELECT * FROM ItemFlavours", conn)

conn.close()

# Add item flavours to Items dataframe
items = items.merge(item_flavours, left_on='ItemFlavourID', right_on='ItemFlavourID', how='left')
items['ItemName'] = items.apply(lambda x: f"{x['ItemName']} ({x['ItemFlavour']})" if pd.notna(x['ItemFlavour']) else x['ItemName'], axis=1)

# Total sales and number of orders
total_sales = orders['OrderAmount'].sum()
total_orders = orders['OrderID'].nunique()

# Sales by payment type
sales_by_payment_type = orders.groupby('PaymentTypeID')['OrderAmount'].sum().reset_index()
sales_by_payment_type = sales_by_payment_type.merge(payment_types, on='PaymentTypeID')
sales_by_payment_type = sales_by_payment_type[sales_by_payment_type['PaymentType'] != 'Coupon']

# Sales by order type
sales_by_order_type = orders.groupby('OrderTypeID')['OrderAmount'].sum().reset_index()
sales_by_order_type = sales_by_order_type.merge(order_types, on='OrderTypeID')

# Most popular items
popular_items = order_items.groupby('ItemID')['ItemAmount'].sum().reset_index()
popular_items = popular_items.merge(items[['ItemID', 'ItemName']], on='ItemID').sort_values(by='ItemAmount', ascending=False)

# Most popular addons (renamed AddOnAmount to AddOnCount)
popular_addons = item_addons.groupby('AddOnID')['AddOnAmount'].sum().reset_index()
popular_addons = popular_addons.merge(addons[['AddOnID', 'AddOn']], on='AddOnID').sort_values(by='AddOnAmount', ascending=False)
popular_addons.rename(columns={'AddOnAmount': 'AddOnCount'}, inplace=True)

# Most popular item and add-on combination
item_addon_combo = item_addons.groupby(['OrderItemID', 'AddOnID'])['AddOnAmount'].sum().reset_index()
item_addon_combo = item_addon_combo.merge(order_items, on='OrderItemID')
item_addon_combo = item_addon_combo.merge(items[['ItemID', 'ItemName']], on='ItemID')
item_addon_combo = item_addon_combo.merge(addons[['AddOnID', 'AddOn']], on='AddOnID')
item_addon_combo = item_addon_combo.groupby(['ItemName', 'AddOn'])['AddOnAmount'].sum().reset_index().sort_values(by='AddOnAmount', ascending=False)
item_addon_combo.rename(columns={'AddOnAmount': 'CombinationCount'}, inplace=True)

# Most popular flavors
popular_flavors = order_items.merge(items[['ItemID', 'ItemFlavour']], on='ItemID')
popular_flavors = popular_flavors.groupby('ItemFlavour')['ItemID'].count().reset_index().rename(columns={'ItemID': 'ItemCount'}).sort_values(by='ItemCount', ascending=False)

# Busiest hours, days, and dates
orders['OrderHour'] = pd.to_datetime(orders['OrderTime']).dt.hour
busiest_hours = orders.groupby('OrderHour')['OrderID'].count().reset_index().rename(columns={'OrderID': 'OrderCount'}).sort_values(by='OrderCount', ascending=False)

orders['OrderDate'] = pd.to_datetime(orders['OrderTime']).dt.date
orders['DayOfWeek'] = pd.to_datetime(orders['OrderDate']).dt.strftime('%A')
busiest_days = orders.groupby('DayOfWeek')['OrderID'].count().reset_index().rename(columns={'OrderID': 'OrderCount'}).sort_values(by='OrderCount', ascending=False)

busiest_actual_day = orders.groupby('OrderDate')['OrderID'].count().reset_index().rename(columns={'OrderID': 'OrderCount'}).sort_values(by='OrderCount', ascending=False)

# Create FullName column by combining CashierName and CashierSurname
cashiers['FullName'] = cashiers['CashierName'] + ' ' + cashiers['CashierSurname']
# Cashier performance metrics
cashier_sales = order_items.merge(orders, on='OrderID').groupby('CashierID')['ItemAmount'].sum().reset_index().merge(cashiers, on='CashierID').sort_values(by='ItemAmount', ascending=False)
cashier_sales.rename(columns={'ItemAmount': 'ItemCount'}, inplace=True)
cashier_traffic = orders.groupby('CashierID')['OrderID'].count().reset_index().rename(columns={'OrderID': 'OrderCount'}).merge(cashiers, on='CashierID').sort_values(by='OrderCount', ascending=False)
cashier_revenue = orders.groupby('CashierID')['OrderAmount'].sum().reset_index().merge(cashiers, on='CashierID').sort_values(by='OrderAmount', ascending=False)

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Function to write DataFrame to Excel with renaming and euro sign formatting
def write_df_to_sheet(df, sheet, start_row=1, start_col=1, rename_columns=None, format_columns=None):
    if rename_columns:
        df = df.rename(columns=rename_columns)
    if format_columns is None:
        format_columns = [col for col in df.columns if 'Amount' in col]
    euro_col_indices = [df.columns.get_loc(col) + start_col for col in format_columns if col in df.columns]
    # Write data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if r_idx > start_row - 1 and c_idx in euro_col_indices:
                cell.number_format = u'€#,##0.00'
    # Adjust column widths considering both the header and data rows
    for c_idx, column in enumerate(df.columns, start_col):
        max_length = 0
        # Consider both the header and data rows
        column_values = [str(value) for value in [df.columns[c_idx - start_col]] + df.iloc[:, c_idx - start_col].tolist()]
        # Find the maximum length of all values in the column (header + data)
        for value_str in column_values:
            max_length = max(max_length, len(value_str))
        # Add a little padding for readability
        adjusted_width = max_length + 2
        column_letter = chr(65 + c_idx - 1)
        sheet.column_dimensions[column_letter].width = adjusted_width

wb = Workbook()
ws = wb.active

ws.title = "Summary"
write_df_to_sheet(pd.DataFrame({'Total Sales Amount': [total_sales], 'Total Order Count': [total_orders]}), ws)

# Write data to each sheet with specified renaming and euro formatting
ws_payment_type = wb.create_sheet("Sales by Payment Type")
write_df_to_sheet(sales_by_payment_type, ws_payment_type)

ws_order_type = wb.create_sheet("Sales by Order Type")
write_df_to_sheet(sales_by_order_type, ws_order_type)

# Use the correct sheet for Popular Items with ItemCount
ws_popular_items = wb.create_sheet("Popular Items")
write_df_to_sheet(popular_items[['ItemName', 'ItemAmount']], ws_popular_items, rename_columns={'ItemAmount': 'ItemCount'})

# Create a new sheet for popular addons (renamed AddOnAmount to AddOnCount)
ws_popular_addons = wb.create_sheet("Popular AddOns")
write_df_to_sheet(popular_addons[['AddOn', 'AddOnCount']], ws_popular_addons)

# Define and use the create_and_insert_top10_chart function to generate and insert the charts
def create_and_insert_top10_chart(data, title, xlabel, ylabel, name_column, count_column, image_path, sheet, cell_position):
    # Sort data to get the top 10
    top10_data = data.nlargest(10, count_column)
    plt.figure(figsize=(10, 6))
    plt.bar(top10_data[name_column], top10_data[count_column])
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.xticks(rotation=45, ha="right")  # Rotate x-axis labels for better readability
    plt.tight_layout()
    plt.savefig(image_path)
    plt.close()
    
    img = Image(image_path)
    sheet.add_image(img, cell_position)

# Create and insert top 10 charts for popular items and addons
create_and_insert_top10_chart(
    popular_items[['ItemName', 'ItemAmount']], 
    'Top 10 Most Popular Items', 
    'Item', 
    'Count', 
    name_column='ItemName', 
    count_column='ItemAmount', 
    image_path=f'{output_dir}/top10_popular_items.png', 
    sheet=ws_popular_items, 
    cell_position='E1'
)

create_and_insert_top10_chart(
    popular_addons[['AddOn', 'AddOnCount']], 
    'Top 10 Most Popular AddOns', 
    'AddOn', 
    'Count', 
    name_column='AddOn', 
    count_column='AddOnCount', 
    image_path=f'{output_dir}/top10_popular_addons.png', 
    sheet=ws_popular_addons, 
    cell_position='E1'
)

# Create and insert top 10 chart for Item-AddOn combos
ws_item_addon_combo = wb.create_sheet("Item-AddOn Combos")
write_df_to_sheet(item_addon_combo[['ItemName', 'AddOn', 'CombinationCount']], ws_item_addon_combo)

# Merge the ItemName and AddOn for the combo chart
item_addon_combo['ItemAddOnCombo'] = item_addon_combo['ItemName'] + ' & ' + item_addon_combo['AddOn']

create_and_insert_top10_chart(
    item_addon_combo[['ItemAddOnCombo', 'CombinationCount']], 
    'Top 10 Item-AddOn Combos', 
    'Item-AddOn Combo', 
    'Count', 
    name_column='ItemAddOnCombo', 
    count_column='CombinationCount', 
    image_path=f'{output_dir}/top10_item_addon_combos.png', 
    sheet=ws_item_addon_combo, 
    cell_position='E1'
)

# Add Busiest Days (Hours, Days, and Weekdays) sheet 
ws_busiest_times = wb.create_sheet("Busiest Times")
write_df_to_sheet(busiest_hours[['OrderHour', 'OrderCount']], ws_busiest_times, start_row=1, start_col=1)
write_df_to_sheet(busiest_days[['DayOfWeek', 'OrderCount']], ws_busiest_times, start_row=1, start_col=4)
write_df_to_sheet(busiest_actual_day[['OrderDate', 'OrderCount']], ws_busiest_times, start_row=1, start_col=7)
# Define the order of days for categorization
day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
# Function to round order times to the nearest half-hour interval
def round_to_nearest_half_hour(order_time):
    hour = order_time.hour
    minute = order_time.minute
    # Round to nearest half-hour mark
    if minute < 15:
        return f"{hour}:00"
    elif minute < 45:
        return f"{hour}:30"
    else:
        return f"{hour + 1}:00" if hour < 22 else "22:00"
# Apply the rounding function to create a new column with half-hour intervals
orders['OrderHalfHour'] = orders['OrderTime'].apply(lambda x: round_to_nearest_half_hour(pd.to_datetime(x)))
# Set DayOfWeek as a categorical variable with the specified order
orders['DayOfWeek'] = pd.Categorical(orders['DayOfWeek'], categories=day_order, ordered=True)
# Group by DayOfWeek and the new OrderHalfHour for average order counts
avg_orders_per_half_hour = orders.groupby(['DayOfWeek', 'OrderHalfHour']).agg({'OrderID': 'count'}).reset_index()
avg_orders_per_half_hour = avg_orders_per_half_hour.rename(columns={'OrderID': 'OrderCount'})
# Pivot table with weekdays as columns and half-hour intervals as rows
avg_orders_per_half_hour_pivot = avg_orders_per_half_hour.pivot_table(
    index='OrderHalfHour', columns='DayOfWeek', values='OrderCount', aggfunc='mean', fill_value=0
)
# Sort the index to ensure proper order for half-hour intervals (12:00 to 22:00)
avg_orders_per_half_hour_pivot = avg_orders_per_half_hour_pivot.reindex([
    f"{hour}:{minute:02d}" for hour in range(12, 23) for minute in (0, 30)
])
# Plot the graph with ordered days in the legend
plt.figure(figsize=(12, 6))
avg_orders_per_half_hour_pivot.plot(kind='line', marker='o', linewidth=2)
plt.title('Average Orders Across Hours and Weekdays')
plt.xlabel('Time of Day')
plt.ylabel('Average Number of Orders')
# Set only whole-hour ticks on the x-axis, from 12:00 to 22:00
whole_hours = [f"{hour}:00" for hour in range(12, 23)]
plt.xticks(ticks=range(0, len(avg_orders_per_half_hour_pivot), 2), labels=whole_hours, rotation=45, ha="right")
plt.legend(title='Weekdays', bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()
# Save the plot as an image
avg_orders_per_half_hour_image_path = f'{output_dir}/avg_orders_per_half_hour_across_weekdays_working_hours.png'
plt.savefig(avg_orders_per_half_hour_image_path)
plt.close()
# Add the image to the Busiest Times sheet at the designated cell
img = Image(avg_orders_per_half_hour_image_path)
ws_busiest_times.add_image(img, 'J1')  



# Add Cashier Performance (Sales, Traffic, and Revenue) sheet 
ws_cashier_performance = wb.create_sheet("Cashier Performance")
write_df_to_sheet(cashier_sales[['FullName', 'ItemCount']], ws_cashier_performance)
write_df_to_sheet(cashier_traffic[['FullName', 'OrderCount']], ws_cashier_performance, start_row=len(cashier_sales) + 3)
write_df_to_sheet(cashier_revenue[['FullName', 'OrderAmount']], ws_cashier_performance, start_row=len(cashier_sales) + len(cashier_traffic) + 6)

# Save the workbook
wb.save(f'{output_dir}/analyzed_data.xlsx')